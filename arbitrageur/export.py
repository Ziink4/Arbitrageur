import json
from math import ceil
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from arbitrageur.crafting import ProfitableItem, profit_per_item, profit_on_cost, profit_per_crafting_step
from arbitrageur.items import Item
from arbitrageur.recipes import Recipe, ITEM_STACK_SIZE

import csv
from logzero import logger


def format_disciplines(disciplines: List[str]) -> str:
    # Returns the first letter of each discipline, except for Armorsmith (Am) & Artificer (At)
    return "/".join([e[0] if e[0] != 'A' else e[:3:2] for e in disciplines])


def format_json_recipe(profitable_item, items_map):
    logger.debug(
        f"""Shopping list for {profitable_item.count} x {items_map[profitable_item.id].name} = {profitable_item.profit} profit ({profit_per_crafting_step(profitable_item)} / step) :""")

    recipe = {}
    for ingredient_id, purchased_ingredient in profitable_item.purchased_ingredients.items():
        ingredient_count = ceil(purchased_ingredient.count)
        if ingredient_count < ITEM_STACK_SIZE:
            ingredient_count_msg = str(ingredient_count)
        else:
            stack_count = ingredient_count // ITEM_STACK_SIZE
            remainder = ingredient_count % ITEM_STACK_SIZE
            if remainder != 0:
                remainder_msg = f""" + {remainder}"""
            else:
                remainder_msg = ""

            ingredient_count_msg = f"""{ingredient_count} ({stack_count} x {ITEM_STACK_SIZE}{remainder_msg})"""

        ingredient_name = items_map[ingredient_id].name
        logger.debug(
            f"""{ingredient_count_msg} {ingredient_name} ({ingredient_id}) for {purchased_ingredient.cost}""")

        recipe[ingredient_name] = {"count": ingredient_count_msg,
                                   "listings": purchased_ingredient.listings}
    return recipe


def generate_export_rows(items_map, profitable_items, recipes_map):
    profitable_items = [e for e in profitable_items if e.profit != 0]
    data = []
    for profitable_item in profitable_items:
        item_id = profitable_item.id
        item = items_map[item_id]
        recipe = recipes_map[item_id]

        item_data = {
            'id':                      item_id,
            'name':                    item.name,
            'rarity':                  item.rarity,
            'disciplines':             format_disciplines(recipe.disciplines),
            'profit':                  profitable_item.profit,
            'crafting_cost':           profitable_item.crafting_cost,
            'count':                   profitable_item.count,
            'avg_profit_per_item':     profit_per_item(profitable_item),
            'roi':                     float(profit_on_cost(profitable_item)),
            'link':                    f"""=HYPERLINK("https://www.gw2bltc.com/en/item/{item.id}")""",
            # TODO: This returns the "total" minimal sell price instead of the minimal sell price per item
            'profitability_threshold': profitable_item.profitability_threshold,
            'time_gated':              profitable_item.time_gated,
            'needs_ascended':          profitable_item.needs_ascended,
            'craft_level':             recipe.min_rating,
            'recipe':                  json.dumps(format_json_recipe(profitable_item, items_map), indent=2)
        }

        data.append(item_data)
    return data


def export_csv(profitable_items: List[ProfitableItem],
               items_map: Dict[int, Item],
               recipes_map: Dict[int, Recipe]) -> None:
    logger.info("Exporting profitable items as CSV")
    data = generate_export_rows(items_map, profitable_items, recipes_map)
    if len(data) == 0:
        logger.warning("Could not find any profitable item to export")
        return

    with open('export.csv', 'w', newline='') as csvfile:
        datawriter = csv.writer(csvfile, delimiter=',', quotechar='"')
        datawriter.writerow(data[0].keys())

        for item_data in data:
            datawriter.writerow(item_data.values())


def get_series_letter(name: str, data: List[dict]) -> str:
    return get_column_letter(list(data[0].keys()).index(name) + 1)


def as_text(value):
    if value is None:
        return ""
    return str(value)


def export_excel(profitable_items: List[ProfitableItem],
                 items_map: Dict[int, Item],
                 recipes_map: Dict[int, Recipe],
                 time_gated: bool = False,
                 needs_ascended: bool = False) -> None:
    logger.info("Exporting profitable items as Excel spreadsheet")
    data = generate_export_rows(items_map, profitable_items, recipes_map)
    if len(data) == 0:
        logger.warning("Could not find any profitable item to export")
        return

    wb = Workbook()
    ws = wb.active

    # add column headings. NB. these must be strings
    ws.append(list(data[0].keys()))
    for row in data:
        if (time_gated or not row['time_gated']) and (needs_ascended or not row['needs_ascended']):
            ws.append(list(row.values()))

    tab = Table(displayName="Data", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    ws.add_table(tab)

    # Applies formatting (format can only be applied to single cells)
    for col_name, number_format in {'profit':                  r"#\,##\,##",
                                    'crafting_cost':           r"#\,##\,##",
                                    'avg_profit_per_item':     r"#\,##\,##",
                                    'roi':                     '0.00%',
                                    'profitability_threshold': r"#\,##\,##"}.items():
        col_letter = get_series_letter(col_name, data)

        for i in range(1, ws.max_row + 1):
            ws[f"{col_letter}{i}"].number_format = number_format

    # Adjusts column width
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length * 0.8 + 5

    wb.save("export.xlsx")
